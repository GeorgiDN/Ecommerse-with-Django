# views.py
from django.http import HttpResponse
from django.shortcuts import render, redirect
from openpyxl import load_workbook
from django.contrib import messages

from ecommerseApp.store.models import ProductOptionValue, ProductOption, ProductVariant, Product
from ecommerseApp.store_admin.forms import ProductImportForm


def import_products(request):
    if not request.user.is_staff:
        return HttpResponse("Unauthorized", status=401)

    if request.method == 'POST':
        form = ProductImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                wb = load_workbook(request.FILES['file'])
                ws = wb.active
                overwrite = form.cleaned_data['overwrite']

                headers = [cell.value for cell in ws[1]]

                # Create mapping of column positions to field names
                column_map = {header: idx for idx, header in enumerate(headers) if header}
                product = None
                # Process rows
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Skip empty rows
                    if not row[column_map.get('name')]:
                        continue

                    # Prepare product data (excluding ID)
                    product_data = {
                        'url_slug': row[column_map.get('url_slug')],
                        'name': row[column_map.get('name')],
                        'description': row[column_map.get('description')],
                        'meta_title': row[column_map.get('meta_title')],
                        'meta_description': row[column_map.get('meta_description')],
                        'model': row[column_map.get('model')],
                        'sku': row[column_map.get('sku')],
                        'tags': row[column_map.get('tags')],
                        'price': row[column_map.get('price')] or 0,
                        'is_on_sale': bool(row[column_map.get('is_on_sale')]),
                        'sale_price': row[column_map.get('sale_price')] or 0,
                        'is_available': bool(row[column_map.get('is_available')]),
                        'is_active': bool(row[column_map.get('is_active')]),
                        'track_quantity': bool(row[column_map.get('track_quantity')]),
                        'quantity': row[column_map.get('quantity')] or 0,
                        'weight': row[column_map.get('weight')] or 0,
                        'has_options': bool(row[column_map.get('has_options')]),
                    }

                    # Handle image separately if needed
                    image_url = row[column_map.get('image')]
                    if image_url and image_url.startswith('http'):
                        # Implement your image download logic here if needed
                        pass

                    # Handle product creation/update
                    if overwrite and 'sku' in product_data and product_data['sku']:
                        product, created = Product.objects.update_or_create(
                            sku=product_data['sku'],
                            defaults=product_data
                        )
                    else:
                        if product is None or (product and product.sku != product_data['sku']):
                            product = Product.objects.create(**product_data)

                    # Process categories
                    if 'category_ids' in column_map and row[column_map['category_ids']]:
                        category_ids = [int(id.strip()) for id in row[column_map['category_ids']].split(',') if id.strip()]
                        product.categories.set(category_ids)

                    # Process options
                    for i in range(1, 4):  # For 3 possible options
                        opt_name = row[column_map.get(f'Option {i} Name')]
                        if opt_name:
                            option, _ = ProductOption.objects.get_or_create(
                                product=product,
                                name=opt_name.strip()
                            )
                            # Process option values
                            opt_values = row[column_map.get(f'Option {i} Values')]
                            if opt_values:
                                for val in opt_values.split(','):
                                    val = val.strip()
                                    if val:
                                        ProductOptionValue.objects.get_or_create(
                                            option=option,
                                            value=val
                                        )

                    # Process variants (look for rows with variant data)
                    variant_sku = row[column_map.get('Variant SKU')]
                    if variant_sku:
                        variant_data = {
                            'product': product,
                            'sku': variant_sku,
                            'price': row[column_map.get('Variant Price')] or 0,
                            'is_on_sale': bool(row[column_map.get('Variant is_on_sale')]),
                            'sale_price': row[column_map.get('Variant sale_price')] or 0,
                            'track_quantity': bool(row[column_map.get('Variant track_quantity')]),
                            'quantity': row[column_map.get('Variant quantity')] or 0,
                            'weight': row[column_map.get('Variant weight')] or 0,
                            'is_available': bool(row[column_map.get('Variant is_available')]),
                        }

                        variant, created = ProductVariant.objects.update_or_create(
                            sku=variant_sku,
                            defaults=variant_data
                        )

                        # Link option values to variant
                        variant_options = row[column_map.get('Variant Options')]
                        if variant_options:
                            for opt_pair in variant_options.split('/'):
                                if ':' in opt_pair:
                                    opt_name, opt_value = [x.strip() for x in opt_pair.split(':')]
                                    try:
                                        option = ProductOption.objects.get(
                                            product=product,
                                            name=opt_name
                                        )
                                        opt_value = ProductOptionValue.objects.get(
                                            option=option,
                                            value=opt_value
                                        )
                                        variant.option_values.add(opt_value)
                                    except (ProductOption.DoesNotExist, ProductOptionValue.DoesNotExist):
                                        continue

                messages.success(request, 'Products imported successfully!')
                return redirect('admin-products')

            except Exception as e:
                messages.error(request, f'Error during import: {str(e)}')
    else:
        form = ProductImportForm()

    return render(request, 'store_admin/admin_products/import_products.html', {'form': form})
