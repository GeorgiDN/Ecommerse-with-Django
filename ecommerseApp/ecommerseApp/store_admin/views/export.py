import csv
from django.contrib import messages
from django.http import HttpResponse
from ecommerseApp.store.models import Product
from django.shortcuts import redirect
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
import csv

from django.http import HttpResponse
from openpyxl import Workbook
import csv



####
def export_products_full(request):
    if not request.user.is_staff:
        return HttpResponse("Unauthorized", status=401)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="products_full_export.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        'id', 'url_slug', 'name', 'description', 'meta_title',
        'meta_description', 'model', 'sku', 'tags', 'price',
        'is_on_sale', 'sale_price', 'image', 'is_available',
        'is_active', 'track_quantity', 'quantity', 'weight',
        'has_options', 'category_names', 'category_ids',
        'Option 1 Name', 'Option 1 Values',
        'Option 2 Name', 'Option 2 Values',
        'Option 3 Name', 'Option 3 Values',
        'Variant SKU', 'Variant Price', 'Variant is_on_sale',
        'Variant sale_price', 'Variant track_quantity', 'Variant quantity',
        'Variant weight', 'Variant is_available',
        'Variant Options',
    ]
    ws.append(headers)

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Fetch all data with optimizations
    products = Product.objects.prefetch_related(
        'categories',
        'product_options',
        'product_variants',
    ).all()

    for product in products:
        category_names = ', '.join([cat.name for cat in product.categories.all()])
        category_ids = ', '.join([str(cat.id) for cat in product.categories.all()])
        base_data = [
            product.id, product.url_slug, product.name, product.description, product.meta_title,
            product.meta_description, product.model, product.sku, product.tags, product.price,
            product.is_on_sale,
            product.sale_price if product.sale_price else 0,
            product.image.url if product.image else '',
            product.is_available, product.is_active, product.track_quantity, product.quantity,
            product.weight, product.has_options, category_names, category_ids,
        ]

        # Get options (max 3 for this example)
        options = product.product_options.all()[:3]
        option_data = []
        for opt in options:
            option_data.extend([
                opt.name,
                ", ".join(v.value for v in opt.option_values.all())
            ])
        # Pad with empty if less than 3 options
        option_data.extend([""] * (6 - len(option_data)))

        # Write product row without variants first
        ws.append(base_data + option_data + [""] * 3)

        # Add variant rows if they exist
        if product.product_variants.exists():
            for variant in product.product_variants.all():
                variant_data = [
                    variant.sku, variant.price, variant.is_on_sale,
                    variant.sale_price, variant.track_quantity,
                    variant.quantity, variant.weight, variant.is_available,
                    " / ".join(
                        f"{ov.option.name}: {ov.value}"
                        for ov in variant.option_values.all()
                    )
                ]
                ws.append(base_data + option_data + variant_data)

    wb.save(response)
    return response


# export products without options
# def products_export(request, format_type='csv'):
#     if not request.user.is_staff:
#         messages.error(request, "You don't have permission to access this.")
#         return redirect('home')
#
#     # Common data preparation
#     products = Product.objects.all()
#     headers = [
#         'id', 'url_slug', 'name', 'description', 'meta_title',
#         'meta_description', 'model', 'sku', 'tags', 'price',
#         'is_on_sale', 'sale_price', 'image', 'is_available',
#         'is_active', 'track_quantity', 'quantity', 'weight',
#         'has_options', 'category_names', 'category_ids'
#     ]
#
#     if format_type == 'excel':
#         return _export_excel(products, headers)
#     else:  # Default to CSV
#         return _export_csv(products, headers)
#
#
# def _export_csv(products, headers):
#     response = HttpResponse(content_type='text/csv')
#     response['Content-Disposition'] = 'attachment; filename="products.csv"'
#
#     writer = csv.writer(response)
#     writer.writerow(headers)
#
#     for product in products:
#         category_names = ', '.join([cat.name for cat in product.categories.all()])
#         category_ids = ', '.join([str(cat.id) for cat in product.categories.all()])
#         writer.writerow([
#             product.id,
#             product.url_slug,
#             product.name,
#             product.description,
#             product.meta_title,
#             product.meta_description,
#             product.model,
#             product.sku,
#             product.tags,
#             product.price,
#             product.is_on_sale,
#             product.sale_price if product.sale_price else 0,
#             product.image.url if product.image else '',
#             product.is_available,
#             product.is_active,
#             product.track_quantity,
#             product.quantity,
#             product.weight,
#             product.has_options,
#             category_names,
#             category_ids,
#         ])
#
#     return response
#
#
# def _export_excel(products, headers):
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
#
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Products"
#
#     # Write headers
#     ws.append(headers)
#
#     # Write data
#     for product in products:
#         category_names = ', '.join([cat.name for cat in product.categories.all()])
#         category_ids = ', '.join([str(cat.id) for cat in product.categories.all()])
#         ws.append([
#             product.id,
#             product.url_slug,
#             product.name,
#             product.description,
#             product.meta_title,
#             product.meta_description,
#             product.model,
#             product.sku,
#             product.tags,
#             product.price,
#             product.is_on_sale,
#             product.sale_price if product.sale_price else 0,
#             product.image.url if product.image else '',
#             product.is_available,
#             product.is_active,
#             product.track_quantity,
#             product.quantity,
#             product.weight,
#             product.has_options,
#             category_names,
#             category_ids,
#         ])
#
#     wb.save(response)
#     return response

